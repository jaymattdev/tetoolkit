"""
Interactive Excel Report Generator - Refactored
Creates user-friendly Excel reports optimized for data comparison and review.
"""

import pandas as pd
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Constants
QUALITY_ORDER = {'Conflict': 0, 'Review': 1, 'Missing': 2, 'Good': 3}

QUALITY_FILLS = {
    'good': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'yellow': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    'red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'light_red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


class ReportGenerator:
    """Generate comprehensive interactive Excel reports for data review."""

    def __init__(self, sharepoint_base_url: str = None, hyperlink_style: str = None,
                 output_link_extension: str = None):
        """
        Initialize the ReportGenerator.

        Args:
            sharepoint_base_url: Base URL for SharePoint links. If None, reads from master config.
            hyperlink_style: Display style for hyperlinks:
                            'short' - Shows "Open File" as clickable text
                            'full'  - Shows the complete URL in the cell
                            If None, reads from master config.
            output_link_extension: File extension for document links (e.g., '.pdf').
                                  If None, reads from master config.
        """
        # Load from master config if not provided
        from config_loader import get_master_config
        config = get_master_config()

        self.sharepoint_base_url = sharepoint_base_url or config.sharepoint_base_url
        self.hyperlink_style = hyperlink_style or config.hyperlink_style
        self.output_link_extension = output_link_extension or config.output_link_extension

        # Output column settings from master config
        self.include_extraction_order = config.include_extraction_order
        self.include_extraction_position = config.include_extraction_position
        self.include_flags = config.include_flags
        self.include_flag_reasons = config.include_flag_reasons

        logger.info(f"Initialized ReportGenerator (hyperlink_style={self.hyperlink_style})")

    # ====================
    # HELPER METHODS
    # ====================

    def _create_sharepoint_link(self, source: str, filename: str) -> str:
        """Create SharePoint link for a document."""
        if pd.notna(filename) and self.output_link_extension:
            # Apply configured output extension
            if not filename.endswith(self.output_link_extension):
                # Strip existing extension if any and add configured extension
                filename = filename.rsplit('.', 1)[0] + self.output_link_extension
        return f"{self.sharepoint_base_url}{source}/{filename}"

    def _validate_participant_df(self, df: pd.DataFrame, context: str) -> pd.DataFrame:
        """Validate and filter DataFrame for participant-based operations."""
        if df.empty or 'participant_id' not in df.columns:
            logger.warning(f"No participant IDs for {context}")
            return pd.DataFrame()

        df_filtered = df[df['participant_id'].notna()].copy()
        if df_filtered.empty:
            logger.warning(f"No extractions with participant IDs for {context}")

        return df_filtered

    def _get_element_quality(self, element_data: pd.DataFrame) -> tuple[str, str, list]:
        """
        Determine quality and best value for an element.

        Returns:
            tuple: (quality, best_value, unique_values_list)
        """
        values = element_data['cleaned_value'].dropna()
        if values.empty:
            values = element_data['value'].dropna()

        unique_values = list(values.unique())

        if len(unique_values) == 0:
            return 'Missing', '', []
        elif len(unique_values) == 1:
            has_low_conf = (element_data['confidence'] == 'LOW').any()
            has_flags = element_data['flags'].notna().any() and (element_data['flags'] != '').any()
            quality = 'Review' if (has_low_conf or has_flags) else 'Good'
            return quality, unique_values[0], unique_values
        else:
            return 'Conflict', ' | '.join([str(v) for v in unique_values]), unique_values

    def _sort_by_quality(self, df: pd.DataFrame, *other_sort_cols) -> pd.DataFrame:
        """Sort DataFrame by quality priority and other columns."""
        if df.empty or 'Quality' not in df.columns:
            return df

        df['_sort'] = df['Quality'].map(QUALITY_ORDER)
        sort_cols = ['_sort'] + list(other_sort_cols)
        df = df.sort_values(sort_cols).drop('_sort', axis=1)
        return df

    def _prepare_filename_and_link(self, df: pd.DataFrame) -> pd.DataFrame:
        """Strip file extensions and add SharePoint links."""
        df['_original_filename'] = df['filename']
        df['filename'] = df['filename'].apply(
            lambda x: x.rsplit('.', 1)[0] if pd.notna(x) and '.' in x else x
        )

        # Use pdf_filename for links if available, otherwise fall back to original filename
        def get_link_filename(row):
            if 'pdf_filename' in row and pd.notna(row.get('pdf_filename')):
                return row['pdf_filename']
            return row['_original_filename']

        df['Document Link'] = df.apply(
            lambda row: self._create_sharepoint_link(row['source'], get_link_filename(row)),
            axis=1
        )
        return df

    # ====================
    # BEST DATA HELPERS
    # ====================

    def _pick_highest_priority_row(self, group: pd.DataFrame, element: str, source_priority: dict) -> pd.Series:
        """Return the row from the highest priority source for this element."""
        from priority_loader import get_highest_priority_source
        sources_present = group['source'].unique().tolist()
        chosen_source = get_highest_priority_source(sources_present, element, source_priority)
        matching = group[group['source'] == chosen_source]
        return matching.iloc[0]

    def _build_best_data_row(self, participant_id, element: str, row: pd.Series, notes: str) -> dict:
        """Build a single row dict for the Best Data tab."""
        pdf_filename = row.get('pdf_filename')
        link_filename = pdf_filename if pd.notna(pdf_filename) else row.get('filename', '')
        return {
            'Participant ID': participant_id,
            'Element': element,
            'Value': row.get('value'),
            'Cleaned Value': row.get('cleaned_value'),
            'Source': row.get('source'),
            'Document Link': self._create_sharepoint_link(str(row.get('source', '')), link_filename),
            'Notes': notes
        }

    def _compute_best_and_review(self, df: pd.DataFrame, source_priority: dict):
        """
        Core logic: determine best value per participant+element and which rows go to review.

        Cases:
            A - 1 unique value across all sources → Best Data: YES, Review: NO
            B - Multiple values, all from same source → Best Data: OMITTED, Review: YES
            C - Multiple values, from different sources → Best Data: YES (highest priority),
                                                          Review: YES (all conflicting rows)

        Returns:
            best_rows  : list of dicts for Best Data tab
            review_entries : list of (original_df_index, in_best_data_note) tuples
        """
        best_rows = []
        review_entries = []  # (idx, note_string)

        df_valid = df[df['participant_id'].notna()].copy()
        if df_valid.empty:
            return best_rows, review_entries

        for (participant_id, element), group in df_valid.groupby(['participant_id', 'element']):
            # Focus on rows that actually have a value
            with_values = group[group['cleaned_value'].notna()]
            if with_values.empty:
                with_values = group[group['value'].notna()]
            if with_values.empty:
                continue  # No value extracted at all — skip

            # Build source → set of unique values map
            source_values: dict = {}
            for idx, row in with_values.iterrows():
                src = row['source']
                val = row.get('cleaned_value') if pd.notna(row.get('cleaned_value')) else row.get('value')
                if src not in source_values:
                    source_values[src] = {'values': set(), 'indices': []}
                if pd.notna(val):
                    source_values[src]['values'].add(str(val))
                source_values[src]['indices'].append(idx)

            all_unique_vals = set()
            for sv in source_values.values():
                all_unique_vals.update(sv['values'])

            num_sources = len(source_values)

            if len(all_unique_vals) <= 1:
                # Case A: All agree (or only one value found)
                if num_sources == 1:
                    notes = "Extracted from 1 document"
                else:
                    notes = f"Same value extracted from {num_sources} documents"

                chosen_row = self._pick_highest_priority_row(with_values, element, source_priority)
                best_rows.append(self._build_best_data_row(participant_id, element, chosen_row, notes))

            else:
                # Conflict exists — check if same or cross-source
                if num_sources == 1:
                    # Case B: Same-source conflict — omit from Best Data
                    for sv in source_values.values():
                        for idx in sv['indices']:
                            review_entries.append((idx, "No - Same source conflict"))
                else:
                    # Case C: Cross-source conflict — use highest priority source
                    notes = (
                        f"Element found in {num_sources} documents with conflicts. "
                        "Highest priority source chosen"
                    )
                    chosen_row = self._pick_highest_priority_row(with_values, element, source_priority)
                    best_rows.append(self._build_best_data_row(participant_id, element, chosen_row, notes))

                    chosen_source = chosen_row['source']
                    for src, sv in source_values.items():
                        in_best = (src == chosen_source)
                        note = "Yes - Used as Best Value" if in_best else "No - Lower priority source"
                        for idx in sv['indices']:
                            review_entries.append((idx, note))

        return best_rows, review_entries

    def create_best_data_tab(self, df: pd.DataFrame, source_priority: dict) -> pd.DataFrame:
        """
        Create Best Data tab — one row per participant+element with the chosen best value.
        Sorted by Participant ID then Element.
        """
        if df.empty:
            return pd.DataFrame()

        logger.info("Creating Best Data tab")
        best_rows, _ = self._compute_best_and_review(df, source_priority)

        if not best_rows:
            logger.warning("No best data rows generated")
            return pd.DataFrame()

        best_df = pd.DataFrame(best_rows)
        best_df = best_df.sort_values(['Participant ID', 'Element'])
        logger.info(f"Best Data tab ready with {len(best_df)} rows")
        return best_df

    def create_review_tab(self, df: pd.DataFrame, source_priority: dict) -> pd.DataFrame:
        """
        Create Review tab — all conflicting rows (Cases B and C), sorted and
        color-banded by participant. Includes 'In Best Data' column.
        """
        if df.empty:
            return pd.DataFrame()

        logger.info("Creating Review tab")
        _, review_entries = self._compute_best_and_review(df, source_priority)

        if not review_entries:
            logger.info("No conflicting rows — Review tab will be empty")
            return pd.DataFrame()

        # Build index → note map (deduplicate by index, keeping first assignment)
        idx_note_map = {}
        for idx, note in review_entries:
            if idx not in idx_note_map:
                idx_note_map[idx] = note

        review_df = df.loc[list(idx_note_map.keys())].copy()
        review_df['In Best Data'] = review_df.index.map(idx_note_map)

        # Add links
        review_df = self._prepare_filename_and_link(review_df)

        # Select columns
        base_cols = ['participant_id', 'source', 'filename', 'element', 'value', 'cleaned_value']
        columns = [c for c in base_cols if c in review_df.columns]
        columns += ['In Best Data', 'Document Link']
        review_df = review_df[[c for c in columns if c in review_df.columns]]

        # Sort by participant then source
        sort_cols = [c for c in ['participant_id', 'source'] if c in review_df.columns]
        if sort_cols:
            review_df = review_df.sort_values(sort_cols)

        logger.info(f"Review tab ready with {len(review_df)} rows")
        return review_df

    # ====================
    # TAB GENERATION
    # ====================

    def create_all_data_tab(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create unified view of all extracted data."""
        if df.empty:
            return pd.DataFrame()

        logger.info(f"Creating All Extracted Data tab with {len(df)} rows")
        report_df = self._prepare_filename_and_link(df.copy())

        # Select columns dynamically based on master config settings
        base_cols = ['participant_id', 'source', 'filename', 'element', 'value', 'cleaned_value']

        # Build optional columns list based on config settings
        optional_cols = []
        if self.include_extraction_order:
            optional_cols.append('extraction_order')
        if self.include_extraction_position:
            optional_cols.append('extraction_position')
        if self.include_flags:
            optional_cols.append('flags')
        if self.include_flag_reasons:
            optional_cols.append('flag_reasons')

        columns = [c for c in base_cols + optional_cols if c in report_df.columns] + ['Document Link']

        report_df = report_df[columns]

        # Sort by participant_id and source for easier reading
        sort_cols = []
        if 'participant_id' in report_df.columns:
            sort_cols.append('participant_id')
        if 'source' in report_df.columns:
            sort_cols.append('source')
        if sort_cols:
            report_df = report_df.sort_values(sort_cols)

        return report_df

    def create_source_statistics_tab(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create source-level statistics."""
        if df.empty:
            return pd.DataFrame()

        logger.info("Creating Source Statistics tab")
        stats_rows = []

        for source, group in df.groupby('source'):
            # Count missing elements (no value extracted)
            missing = len(group[group['cleaned_value'].isna() & group['value'].isna()])

            # Count actual extractions (exclude missing - only count rows with values)
            actual_extractions = len(group) - missing

            # High/Low confidence only applies to actual extractions
            high_conf = len(group[group.get('confidence', 'HIGH') == 'HIGH']) if 'confidence' in group.columns else actual_extractions
            low_conf = len(group[group.get('confidence', 'LOW') == 'LOW']) if 'confidence' in group.columns else 0

            # Calculate percentages based on total elements (actual + missing)
            total_elements = len(group)
            high_pct = (high_conf / total_elements * 100) if total_elements > 0 else 0
            low_pct = (low_conf / total_elements * 100) if total_elements > 0 else 0
            missing_pct = (missing / total_elements * 100) if total_elements > 0 else 0

            problem_elements = "None"
            if 'confidence' in group.columns:
                problems = group[group['confidence'] == 'LOW']['element'].value_counts().head(3)
                problem_elements = ", ".join([f"{elem} ({count})" for elem, count in problems.items()])

            stats_rows.append({
                'Source': source,
                'Total Extractions': actual_extractions,
                'High Confidence': high_conf,
                'High Conf %': round(high_pct, 1),
                'Low Confidence': low_conf,
                'Low Conf %': round(low_pct, 1),
                'Missing Data': missing,
                'Missing %': round(missing_pct, 1),
                'Top Problem Elements': problem_elements
            })

        stats_df = pd.DataFrame(stats_rows).sort_values('High Conf %', ascending=True)
        logger.info(f"Source Statistics tab ready with {len(stats_df)} sources")
        return stats_df

    def create_participant_statistics_tab(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create participant-level statistics."""
        df_with_id = self._validate_participant_df(df, "participant statistics")
        if df_with_id.empty:
            return pd.DataFrame()

        logger.info("Creating Participant Statistics tab")

        # Count total elements
        all_elements = df_with_id['element'].unique()
        total_elements = len(all_elements)
        stats_rows = []

        for participant_id, group in df_with_id.groupby('participant_id'):
            elements_extracted = group['element'].nunique()
            completeness = (elements_extracted / total_elements * 100) if total_elements > 0 else 0

            # Count conflicts by checking each element for multiple unique values
            conflicts = 0
            for element in all_elements:
                element_data = group[group['element'] == element]
                quality, _, _ = self._get_element_quality(element_data)
                if quality == 'Conflict':
                    conflicts += 1

            stats_rows.append({
                'Participant ID': participant_id,
                'Elements Extracted': elements_extracted,
                'Total Elements': total_elements,
                'Completeness %': round(completeness, 1),
                'Conflicts': conflicts
            })

        stats_df = pd.DataFrame(stats_rows).sort_values('Completeness %', ascending=False)
        logger.info(f"Participant Statistics tab ready with {len(stats_df)} participants")
        return stats_df

    # ====================
    # EXCEL FORMATTING
    # ====================

    def _apply_conditional_formatting_rule(self, ws, col_letter, value, fill, font=None):
        """Helper to apply conditional formatting rule."""
        try:
            rule = CellIsRule(operator='equal', formula=[f'"{value}"'], fill=fill)
            if font:
                rule.font = font
            ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}', rule)
        except Exception as e:
            logger.warning(f"Could not apply conditional formatting for {value} in {col_letter}: {e}")

    def _apply_participant_color_banding(self, ws, headers):
        """Apply alternating grey/white color banding by participant ID."""
        # Find participant_id column
        pid_col = None
        for header, idx in headers.items():
            if header == 'participant_id':
                pid_col = idx
                break

        if not pid_col:
            logger.warning("No participant_id column found for color banding")
            return

        # Define alternating fills
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Track participant changes and apply banding
        current_participant = None
        use_grey = False

        for row_idx in range(2, ws.max_row + 1):
            participant = ws.cell(row_idx, pid_col).value

            # Check if participant changed
            if participant != current_participant:
                current_participant = participant
                use_grey = not use_grey  # Toggle color

            # Apply fill to entire row
            fill = grey_fill if use_grey else white_fill
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                # Only apply if cell doesn't already have special formatting
                if cell.fill.start_color.rgb in ('00000000', None, 'FFFFFF', 'F2F2F2'):
                    cell.fill = fill

        logger.info(f"Applied participant color banding to {ws.max_row - 1} rows")

    def apply_excel_formatting(self, filepath: str):
        """Apply comprehensive Excel formatting to all tabs."""
        try:
            logger.info(f"Applying formatting to {filepath}")
            wb = load_workbook(filepath)

            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            link_font = Font(color="0563C1", underline="single")

            for ws in wb.worksheets:
                if ws.max_row == 1:
                    continue

                # Format headers
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Build column map
                headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
                col_map = {}

                for header, idx in headers.items():
                    col_letter = ws.cell(1, idx).column_letter

                    # Map columns for conditional formatting
                    if header == 'Quality':
                        col_map['quality'] = col_letter

                # Apply conditional formatting
                self._apply_formatting_rules(ws, col_map)

                # Apply color banding by participant for All Extracted Data and Review tabs
                if ws.title in ('All Extracted Data', 'Review'):
                    self._apply_participant_color_banding(ws, headers)

                # Process hyperlinks
                self._process_hyperlinks(ws, link_font)

                # Auto-adjust widths
                self._auto_adjust_columns(ws)

                # Freeze and filter
                ws.freeze_panes = 'A2'
                if ws.max_row > 1:
                    ws.auto_filter.ref = ws.dimensions

            wb.save(filepath)
            logger.info("Formatting applied successfully")

        except Exception as e:
            logger.error(f"Error applying formatting: {e}", exc_info=True)

    def _apply_formatting_rules(self, ws, col_map):
        """Apply conditional formatting rules for Quality column."""
        fills = QUALITY_FILLS

        if 'quality' in col_map:
            self._apply_conditional_formatting_rule(ws, col_map['quality'], 'Good', fills['good'])
            self._apply_conditional_formatting_rule(ws, col_map['quality'], 'Review', fills['yellow'])
            self._apply_conditional_formatting_rule(ws, col_map['quality'], 'Conflict', fills['red'])
            self._apply_conditional_formatting_rule(ws, col_map['quality'], 'Missing', fills['light_red'])

    def _process_hyperlinks(self, ws, link_font):
        """Process and convert hyperlinks, creating hidden columns for additional links."""
        use_short_style = (self.hyperlink_style == 'short')

        # Process regular link columns
        header_row = [cell.value for cell in ws[1]]
        link_columns = [idx for idx, h in enumerate(header_row, 1) if 'Link' in str(h) or 'Documents' in str(h)]

        for col_idx in link_columns:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                cell_value = str(cell.value or '')

                if cell_value and cell_value.startswith('http'):
                    if ' | ' in cell_value:
                        urls = cell_value.split(' | ')
                        cell.hyperlink = urls[0]
                        if use_short_style:
                            cell.value = f"Open File ({len(urls)} docs)"
                        # else: keep full URL as cell value
                    else:
                        cell.hyperlink = cell_value
                        if use_short_style:
                            cell.value = "Open File"
                        # else: keep full URL as cell value
                    cell.font = link_font

        # Track columns that need helper columns
        columns_needing_helpers = {}  # {col_idx: max_files_needed}

        # First pass: identify columns with multiple files
        for col_idx in range(1, ws.max_column + 1):
            max_files = 0
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                cell_value = str(cell.value or '')

                if '|||' in cell_value and ' | ' in cell_value:
                    file_count = cell_value.count(' | ') + 1
                    max_files = max(max_files, file_count)

            if max_files > 1:
                columns_needing_helpers[col_idx] = max_files

        # Insert helper columns (right to left to maintain indices)
        for col_idx in sorted(columns_needing_helpers.keys(), reverse=True):
            max_files = columns_needing_helpers[col_idx]
            header_name = ws.cell(1, col_idx).value

            # Insert helper columns after this column
            for helper_idx in range(max_files - 1):
                ws.insert_cols(col_idx + 1)
                helper_col = col_idx + 1 + helper_idx
                ws.cell(1, helper_col).value = f"{header_name}_Link{helper_idx + 2}"
                ws.column_dimensions[get_column_letter(helper_col)].hidden = True

        # Second pass: process hyperlinks and populate helper columns
        for col_idx in range(1, ws.max_column + 1):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                cell_value = str(cell.value or '')

                if '|||' in cell_value:
                    # Check if this has multiple pipe-separated entries
                    if ' | ' in cell_value:
                        # Multiple files: "file1.pdf|||link1 | file2.pdf|||link2"
                        parts = cell_value.split(' | ')
                        display_parts = []
                        links = []

                        for part in parts:
                            if '|||' in part:
                                value, link = part.split('|||', 1)
                                display_parts.append(value)
                                if link and link.startswith('http'):
                                    links.append(link)

                        # Main cell: show filenames or short text based on style
                        if use_short_style:
                            cell.value = f"Open File ({len(display_parts)} docs)"
                        else:
                            cell.value = ' | '.join(display_parts)

                        if links:
                            cell.hyperlink = links[0]
                            cell.font = link_font

                        # Populate helper columns with additional links
                        for helper_idx, link in enumerate(links[1:], start=1):
                            helper_cell = ws.cell(row_idx, col_idx + helper_idx)
                            if use_short_style:
                                helper_cell.value = f"Open File {helper_idx + 1}"
                            else:
                                helper_cell.value = display_parts[helper_idx] if helper_idx < len(display_parts) else ''
                            helper_cell.hyperlink = link
                            helper_cell.font = link_font
                    else:
                        # Single file: "file.pdf|||link"
                        value, link = cell_value.split('|||', 1)
                        if link and link.startswith('http'):
                            if use_short_style:
                                cell.value = "Open File"
                            else:
                                cell.value = value
                            cell.hyperlink = link
                            cell.font = link_font

    def _auto_adjust_columns(self, ws, sample_size=100):
        """Auto-adjust column widths based on content."""
        for col_idx in range(1, ws.max_column + 1):
            col_letter = ws.cell(1, col_idx).column_letter
            max_length = len(str(ws.cell(1, col_idx).value or ''))

            for row_idx in range(2, min(sample_size + 2, ws.max_row + 1)):
                try:
                    cell_value = str(ws.cell(row_idx, col_idx).value or '')
                    max_length = max(max_length, len(cell_value))
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # ====================
    # MAIN REPORT GENERATION
    # ====================

    def generate_interactive_report(
        self,
        validated_df: pd.DataFrame,
        output_path: str,
        plan_name: str,
        source_priority: dict = None
    ) -> str:
        """
        Generate comprehensive interactive Excel report.

        Args:
            validated_df: Full validated extractions DataFrame
            output_path: Output file path for the Excel report
            plan_name: Plan name for logging
            source_priority: Dict mapping element -> ordered list of sources (highest first).
                            Load with priority_loader.load_source_priority(). If None or empty,
                            falls back to first available source for conflicts.
        """
        source_priority = source_priority or {}
        logger.info(f"Generating interactive report for {plan_name}")

        # Compute best data and review together (single pass through data)
        best_rows, review_entries = self._compute_best_and_review(validated_df, source_priority)

        # Build each tab
        best_data = pd.DataFrame(best_rows).sort_values(['Participant ID', 'Element']) if best_rows else pd.DataFrame()

        # Review tab needs separate build (uses review_entries from above)
        if review_entries:
            idx_note_map = {}
            for idx, note in review_entries:
                if idx not in idx_note_map:
                    idx_note_map[idx] = note
            review_df = validated_df.loc[list(idx_note_map.keys())].copy()
            review_df['In Best Data'] = review_df.index.map(idx_note_map)
            review_df = self._prepare_filename_and_link(review_df)
            base_cols = ['participant_id', 'source', 'filename', 'element', 'value', 'cleaned_value']
            review_cols = [c for c in base_cols if c in review_df.columns] + ['In Best Data', 'Document Link']
            review_df = review_df[[c for c in review_cols if c in review_df.columns]]
            sort_cols = [c for c in ['participant_id', 'source'] if c in review_df.columns]
            if sort_cols:
                review_df = review_df.sort_values(sort_cols)
        else:
            review_df = pd.DataFrame()

        all_data = self.create_all_data_tab(validated_df)
        participant_stats = self.create_participant_statistics_tab(validated_df)
        source_stats = self.create_source_statistics_tab(validated_df)

        # Write to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            tabs = [
                (best_data, 'Best Data'),
                (all_data, 'All Extracted Data'),
                (review_df, 'Review'),
                (participant_stats, 'Participant Statistics'),
                (source_stats, 'Source Statistics'),
            ]
            for df, sheet_name in tabs:
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"Added {sheet_name}: {len(df)} rows")

        # Apply formatting
        self.apply_excel_formatting(output_path)

        logger.info(f"Interactive report generated: {output_path}")
        return output_path


def main():
    """Example usage."""
    pass


if __name__ == "__main__":
    main()
